"""
reporter.py
分析結果をCSV・Excelファイルとして出力するモジュール
"""
from __future__ import annotations

import csv
import json
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import TYPE_CHECKING

import openpyxl
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter
from loguru import logger

from .analyzer import ArticleStructure
from .keyword_extractor import KeywordResult
from .seo_checker import SeoResult


# ── スタイル定数 ─────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", fgColor="2D3E50")
HEADER_FONT   = Font(color="FFFFFF", bold=True, size=11)
ACCENT_FILL   = PatternFill("solid", fgColor="E8F4FD")
GOOD_FILL     = PatternFill("solid", fgColor="D4EDDA")
BAD_FILL      = PatternFill("solid", fgColor="F8D7DA")
BORDER_SIDE   = Side(style="thin", color="CCCCCC")
CELL_BORDER   = Border(left=BORDER_SIDE, right=BORDER_SIDE,
                        top=BORDER_SIDE, bottom=BORDER_SIDE)
CENTER_ALIGN  = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN    = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _style_header_row(ws, row: int, cols: int):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = CELL_BORDER


def _auto_width(ws, min_width=12, max_width=50):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)


class Reporter:
    def __init__(self, output_dir: str = "output", config: dict | None = None):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.config = config or {}

    def generate(
        self,
        structures: list[ArticleStructure],
        seo_results: list[SeoResult],
        keyword_results: list[KeywordResult],
    ) -> dict[str, Path]:
        """
        全結果をまとめてCSV・Excelに出力する
        Returns: {"excel": Path, "csv_articles": Path, ...}
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        prefix = self.config.get("filename_prefix", "competitor_analysis")
        output_files = {}

        # 記事レベルのデータを統合
        rows = self._merge_rows(structures, seo_results, keyword_results)

        # ── CSV出力 ────────────────────────────────────────────
        if self.config.get("csv", True):
            csv_path = self.output_dir / f"{prefix}_articles_{timestamp}.csv"
            self._write_articles_csv(rows, csv_path)
            output_files["csv_articles"] = csv_path

        # ── Excel出力 ──────────────────────────────────────────
        if self.config.get("excel", True):
            xlsx_path = self.output_dir / f"{prefix}_{timestamp}.xlsx"
            self._write_excel(rows, structures, seo_results, keyword_results, xlsx_path)
            output_files["excel"] = xlsx_path

        return output_files

    # ─────────────────────────────────────────────────────────
    # 内部メソッド
    # ─────────────────────────────────────────────────────────

    def _merge_rows(
        self,
        structures: list[ArticleStructure],
        seo_results: list[SeoResult],
        keyword_results: list[KeywordResult],
    ) -> list[dict]:
        seo_map = {r.url: r for r in seo_results}
        kw_map  = {r.url: r for r in keyword_results}

        rows = []
        for s in structures:
            seo = seo_map.get(s.url, SeoResult(url=s.url))
            kw  = kw_map.get(s.url, KeywordResult(url=s.url, domain=s.domain))

            top_kw_freq  = ", ".join(w for w, _ in kw.keywords_freq[:5])
            top_kw_tfidf = ", ".join(w for w, _ in kw.keywords_tfidf[:5])

            rows.append({
                "url":                  s.url,
                "domain":               s.domain,
                "title":                s.title,
                "word_count":           s.word_count,
                "paragraph_count":      s.paragraph_count,
                "image_count":          s.image_count,
                "h1_count":             s.h1_count,
                "h2_count":             s.h2_count,
                "h3_count":             s.h3_count,
                "h4_count":             s.h4_count,
                "total_headings":       s.total_heading_count,
                "max_heading_depth":    s.max_heading_depth,
                "avg_heading_length":   round(s.avg_heading_length, 1),
                "load_time_ms":         round(s.load_time_ms, 0),
                "status_code":          s.status_code,
                "error":                s.error or "",
                # SEO
                "meta_title":           seo.meta_title,
                "meta_title_length":    seo.meta_title_length,
                "title_ok":             seo.title_ok,
                "meta_description":     seo.meta_description,
                "meta_description_length": seo.meta_description_length,
                "description_ok":       seo.description_ok,
                "has_ogp":              seo.has_ogp,
                "has_og_title":         seo.has_og_title,
                "has_og_description":   seo.has_og_description,
                "has_og_image":         seo.has_og_image,
                "has_canonical":        seo.has_canonical,
                "has_structured_data":  seo.has_structured_data,
                "internal_links":       seo.internal_links,
                "external_links":       seo.external_links,
                "seo_score":            seo.seo_score,
                # Keywords
                "top_keywords_freq":    top_kw_freq,
                "top_keywords_tfidf":   top_kw_tfidf,
            })
        return rows

    # ── CSV ──────────────────────────────────────────────────

    def _write_articles_csv(self, rows: list[dict], path: Path):
        if not rows:
            return
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
            writer.writeheader()
            writer.writerows(rows)
        logger.info(f"CSV出力: {path}")

    # ── Excel ────────────────────────────────────────────────

    def _write_excel(
        self,
        rows: list[dict],
        structures: list[ArticleStructure],
        seo_results: list[SeoResult],
        keyword_results: list[KeywordResult],
        path: Path,
    ):
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # デフォルトシートを削除

        self._sheet_summary(wb, rows, structures, seo_results)
        self._sheet_articles(wb, rows)
        self._sheet_keywords(wb, keyword_results)
        self._sheet_headings(wb, structures)

        wb.save(path)
        logger.info(f"Excel出力: {path}")

    def _sheet_summary(self, wb, rows, structures, seo_results):
        ws = wb.create_sheet("サマリー")

        # ドメイン別に集計
        domain_data: dict[str, list] = defaultdict(list)
        for r in rows:
            domain_data[r["domain"]].append(r)

        headers = [
            "ドメイン", "記事数", "平均文字数", "平均見出し数",
            "平均SEOスコア", "平均ロード時間(ms)", "最頻出キーワード"
        ]
        ws.append(headers)
        _style_header_row(ws, 1, len(headers))

        for domain, items in domain_data.items():
            n = len(items)
            avg_wc    = round(sum(i["word_count"] for i in items) / n, 0)
            avg_hdg   = round(sum(i["total_headings"] for i in items) / n, 1)
            avg_seo   = round(sum(i["seo_score"] for i in items) / n, 1)
            avg_load  = round(sum(i["load_time_ms"] for i in items) / n, 0)

            # ドメインの全キーワードを集約
            kw_counter: Counter = Counter()  # type: ignore
            for i in items:
                for kw in i["top_keywords_freq"].split(", "):
                    if kw:
                        kw_counter[kw] += 1
            top_kw = kw_counter.most_common(1)[0][0] if kw_counter else ""

            ws.append([domain, n, avg_wc, avg_hdg, avg_seo, avg_load, top_kw])

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = CELL_BORDER
                cell.alignment = CENTER_ALIGN
                if cell.row % 2 == 0:
                    cell.fill = ACCENT_FILL

        _auto_width(ws)
        ws.freeze_panes = "A2"

    def _sheet_articles(self, wb, rows: list[dict]):
        ws = wb.create_sheet("記事詳細")

        headers = [
            "URL", "ドメイン", "タイトル", "文字数", "段落数", "画像数",
            "H1", "H2", "H3", "H4", "見出し合計", "見出し最大深さ",
            "SEOスコア", "ロード(ms)", "タイトルOK", "DESC OK",
            "OGP有無", "Canonical", "構造化データ", "内部リンク", "外部リンク",
            "頻度キーワードTOP5", "TF-IDFキーワードTOP5"
        ]
        ws.append(headers)
        _style_header_row(ws, 1, len(headers))

        col_map = {
            "URL": "url", "ドメイン": "domain", "タイトル": "title",
            "文字数": "word_count", "段落数": "paragraph_count", "画像数": "image_count",
            "H1": "h1_count", "H2": "h2_count", "H3": "h3_count", "H4": "h4_count",
            "見出し合計": "total_headings", "見出し最大深さ": "max_heading_depth",
            "SEOスコア": "seo_score", "ロード(ms)": "load_time_ms",
            "タイトルOK": "title_ok", "DESC OK": "description_ok",
            "OGP有無": "has_ogp", "Canonical": "has_canonical",
            "構造化データ": "has_structured_data", "内部リンク": "internal_links",
            "外部リンク": "external_links",
            "頻度キーワードTOP5": "top_keywords_freq", "TF-IDFキーワードTOP5": "top_keywords_tfidf",
        }

        for row_data in rows:
            row_vals = [row_data.get(col_map[h], "") for h in headers]
            ws.append(row_vals)

        # スコアに応じた色付け
        seo_col = headers.index("SEOスコア") + 1
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = CELL_BORDER
                cell.alignment = LEFT_ALIGN
            seo_cell = ws.cell(row=row[0].row, column=seo_col)
            if isinstance(seo_cell.value, (int, float)):
                seo_cell.fill = GOOD_FILL if seo_cell.value >= 70 else (
                    ACCENT_FILL if seo_cell.value >= 40 else BAD_FILL
                )

        _auto_width(ws)
        ws.freeze_panes = "A2"

    def _sheet_keywords(self, wb, keyword_results: list[KeywordResult]):
        ws = wb.create_sheet("キーワード")

        headers = ["URL", "ドメイン", "キーワード", "頻度", "TF-IDFスコア", "種別"]
        ws.append(headers)
        _style_header_row(ws, 1, len(headers))

        for kr in keyword_results:
            # 頻度
            for word, freq in kr.keywords_freq:
                ws.append([kr.url, kr.domain, word, freq, "", "頻度"])
            # TF-IDF
            for word, score in kr.keywords_tfidf:
                ws.append([kr.url, kr.domain, word, "", score, "TF-IDF"])
            # 共起
            for w1, w2, count in kr.cooccurrence_pairs:
                ws.append([kr.url, kr.domain, f"{w1} × {w2}", count, "", "共起"])

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = CELL_BORDER
                cell.alignment = LEFT_ALIGN

        _auto_width(ws)
        ws.freeze_panes = "A2"

    def _sheet_headings(self, wb, structures: list[ArticleStructure]):
        ws = wb.create_sheet("見出し一覧")

        headers = ["URL", "ドメイン", "タイトル", "レベル", "見出しテキスト"]
        ws.append(headers)
        _style_header_row(ws, 1, len(headers))

        level_fills = {
            1: PatternFill("solid", fgColor="2D3E50"),
            2: PatternFill("solid", fgColor="4A6FA5"),
            3: PatternFill("solid", fgColor="7FB3D3"),
            4: PatternFill("solid", fgColor="BDD7EE"),
        }
        level_fonts = {
            1: Font(color="FFFFFF", bold=True),
            2: Font(color="FFFFFF"),
            3: Font(color="000000"),
            4: Font(color="000000"),
        }

        for s in structures:
            for level, texts in [
                (1, s.h1_texts), (2, s.h2_texts),
                (3, s.h3_texts), (4, s.h4_texts)
            ]:
                for text in texts:
                    row_num = ws.max_row + 1
                    ws.append([s.url, s.domain, s.title, f"H{level}", text])
                    for col in range(1, 6):
                        cell = ws.cell(row=row_num, column=col)
                        cell.fill = level_fills[level]
                        cell.font = level_fonts[level]
                        cell.border = CELL_BORDER
                        cell.alignment = LEFT_ALIGN

        _auto_width(ws)
        ws.freeze_panes = "A2"
